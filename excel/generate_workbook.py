# -*- coding: utf-8 -*-
"""
工事現場 工程表アプリ (Excel版) ジェネレーター

1つのブックの中で
  - 工程マスタ   : 作業の入力はここ1か所
  - 全体工程表   : 月単位ガント + マイルストーン + 予定出来高曲線
  - 月間工程表   : 対象年月を選ぶと日単位ガントを自動生成
  - 日割り工程表 : 基準日から14日間の日別詳細
が全て数式で連動する。

実行: python generate_workbook.py
出力: 工事工程表アプリ.xlsx
"""
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

# ---------------------------------------------------------------- 定数・スタイル
FONT = "Yu Gothic"
MASTER_FIRST = 5          # 工程マスタ データ開始行
MASTER_LAST = 154         # 工程マスタ データ最終行 (150作業)
MR = f"$5:${MASTER_LAST}" # よく使う行レンジ表記用

C_HEADER = "44546A"       # 濃紺 (ヘッダー)
C_HEADER2 = "8496B0"      # 淡紺 (サブヘッダー)
C_INPUT = "FFF2CC"        # 入力セル(黄)
C_CALC = "F2F2F2"         # 自動計算セル(灰)
C_PLAN = "9DC3E6"         # 予定バー(水色)
C_DONE = "2F5597"         # 実績バー(濃青)
C_SAT = "DDEBF7"          # 土曜
C_SUN = "FCE4E4"          # 日曜・祝日
C_TODAY = "FFE7E7"        # 当月ハイライト
C_MS = "C00000"           # マイルストーン▼

thin = Side(style="thin", color="BFBFBF")
medium = Side(style="medium", color="44546A")
b_all = Border(left=thin, right=thin, top=thin, bottom=thin)
b_head = Border(left=thin, right=thin, top=medium, bottom=medium)

f_title = Font(name=FONT, size=14, bold=True, color="1F3864")
f_head = Font(name=FONT, size=9, bold=True, color="FFFFFF")
f_base = Font(name=FONT, size=9)
f_small = Font(name=FONT, size=8)
f_ms = Font(name=FONT, size=9, bold=True, color=C_MS)
f_label = Font(name=FONT, size=9, bold=True)

fill_head = PatternFill("solid", fgColor=C_HEADER)
fill_head2 = PatternFill("solid", fgColor=C_HEADER2)
fill_input = PatternFill("solid", fgColor=C_INPUT)
fill_calc = PatternFill("solid", fgColor=C_CALC)

al_c = Alignment(horizontal="center", vertical="center")
al_l = Alignment(horizontal="left", vertical="center")
al_cw = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = Workbook()


def style(ws, ref, font=None, fill=None, border=None, align=None, numfmt=None):
    """レンジ一括スタイル適用"""
    for row in ws[ref]:
        for c in row:
            if font: c.font = font
            if fill: c.fill = fill
            if border: c.border = border
            if align: c.alignment = align
            if numfmt: c.number_format = numfmt


# ================================================================ 1. 使い方
ws = wb.active
ws.title = "使い方"
ws.sheet_properties.tabColor = "808080"
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 95

ws["B2"] = "工事工程表アプリ(Excel版) の使い方"
ws["B2"].font = f_title

rows = [
    ("■ 仕組み", "作業の入力は「工程マスタ」シート1か所だけ。全体工程表(月単位)・月間工程表(日単位)・"
                 "日割り工程表(日単位詳細)は、すべて工程マスタから数式で自動生成されます。"),
    ("", ""),
    ("① 設定", "工事名・工期・全体工程表の表示開始月・祝日リスト・工種リストを入力します。"),
    ("② 工程マスタ", "作業ごとに 工種/作業名/担当業者/開始日/終了日/進捗率 を入力します(最大150行)。"
                     "マイルストーン(着工・上棟・検査・竣工など)は開始日=終了日にして「MS」列に○を入れると、"
                     "工程表上に赤い▼で表示されます。"),
    ("③ 全体工程表", "自動生成。表示開始月は「設定」シートで変更できます(24か月表示)。"
                     "最下部に予定出来高曲線(自動計算)と実績出来高の入力行(黄色)があり、"
                     "グラフに赤=予定/青=実績の2本のS字カーブを表示します。実績は毎月末に累計%を入力してください。"),
    ("④ 月間工程表", "黄色セルで対象の年・月を選ぶと、その月に掛かる作業だけを自動抽出して日単位のガントを表示します。"),
    ("⑤ 日割り工程表", "黄色セルに基準日を入力すると、その日から14日間の日別工程と日ごとの稼働作業数を表示します。"),
    ("⑥ ネットワーク工程表", "工程マスタの「先行1〜3」列に先行作業のNo(自分より小さいNoのみ)を入力すると、"
                     "CPM(クリティカルパス法)で最早開始/完了・最遅開始/完了・余裕日数を自動計算し、"
                     "週単位のタイムラインに自動配置します。赤=クリティカルパス(遅らせると竣工が遅れる作業)、"
                     "青=余裕のある作業、黄=フロート(遅らせられる幅)、数字=作業No。"
                     "先行のない作業は工期(自)を開始日として計算します。不正な先行Noは工程マスタ上で赤く警告されます。"),
    ("", ""),
    ("■ バーの見方", "水色=予定 / 濃青=進捗率に応じた実績(完了分) / 赤▼=マイルストーン / 黄色列=本日"),
    ("■ 土日祝", "月間・日割り工程表では 土曜=青系、日曜・祝日=赤系 に自動で色分けされます。"
                 "祝日は「設定」シートのリストを編集してください。"),
    ("", ""),
    ("■ セルの色", "黄色セル=入力する場所 / 灰色セル=自動計算(触らない) / 白セルの表部分=自動生成"),
    ("■ 注意", "全体工程表の表示は100行、月間・日割り工程表は各50行までです。"
               "行の挿入・削除はせず、工程マスタの空き行に追記してください。"),
]
r = 4
for label, text in rows:
    ws.cell(row=r, column=2, value=label).font = f_label
    c = ws.cell(row=r, column=3, value=text)
    c.font = f_base
    c.alignment = Alignment(vertical="top", wrap_text=True)
    if text and len(text) > 60:
        ws.row_dimensions[r].height = 28
    r += 1

# 凡例サンプル
ws.cell(row=r + 1, column=2, value="凡例:").font = f_label
ws.cell(row=r + 1, column=3, value="  予定バー ")
ws.cell(row=r + 1, column=3).fill = PatternFill("solid", fgColor=C_PLAN)
ws.cell(row=r + 2, column=3, value="  実績バー ")
ws.cell(row=r + 2, column=3).fill = PatternFill("solid", fgColor=C_DONE)
ws.cell(row=r + 2, column=3).font = Font(name=FONT, size=9, color="FFFFFF")

ws.page_setup.orientation = "portrait"
ws.page_setup.paperSize = 9
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws.print_area = "A1:C25"

# ================================================================ 2. 設定
ws = wb.create_sheet("設定")
ws.sheet_properties.tabColor = "A6A6A6"
for col, w in {"A": 16, "B": 26, "C": 3, "D": 16, "E": 14, "F": 3,
               "G": 3, "H": 14, "I": 3, "J": 16}.items():
    ws.column_dimensions[col].width = w

ws["A1"] = "設定"
ws["A1"].font = f_title

labels = [
    ("A3", "工事名", "B3", "(仮称)○○ビル新築工事"),
    ("A4", "発注者", "B4", "○○株式会社"),
    ("A5", "元請", "B5", "△△建設株式会社"),
    ("A6", "現場代理人", "B6", "山田 太郎"),
    ("A8", "工期(自)", "B8", dt.date(2026, 7, 1)),
    ("A9", "工期(至)", "B9", dt.date(2027, 3, 31)),
]
for la, lv, ca, cv in labels:
    ws[la] = lv
    ws[la].font = f_label
    ws[ca] = cv
    ws[ca].font = f_base
    ws[ca].fill = fill_input
    ws[ca].border = b_all
    if isinstance(cv, dt.date):
        ws[ca].number_format = "yyyy/m/d"

ws["A11"] = "全体工程表"
ws["A11"].font = f_label
ws["A12"] = "表示開始月"
ws["A12"].font = f_base
ws["B12"] = "=DATE(YEAR($B$8),MONTH($B$8),1)"
ws["B12"].number_format = "yyyy年m月"
ws["B12"].fill = fill_input
ws["B12"].border = b_all
ws["B12"].font = f_base
ws["A13"] = "(24か月表示)"
ws["A13"].font = f_small

# 祝日リスト
ws["D3"] = "祝日リスト(編集可)"
ws["D3"].font = f_label
ws["D4"] = "日付"
ws["E4"] = "名称"
for a in ("D4", "E4"):
    ws[a].font = f_head
    ws[a].fill = fill_head
    ws[a].border = b_head
    ws[a].alignment = al_c
holidays = [
    ("2026/1/1", "元日"), ("2026/1/12", "成人の日"), ("2026/2/11", "建国記念の日"),
    ("2026/2/23", "天皇誕生日"), ("2026/3/20", "春分の日"), ("2026/4/29", "昭和の日"),
    ("2026/5/3", "憲法記念日"), ("2026/5/4", "みどりの日"), ("2026/5/5", "こどもの日"),
    ("2026/5/6", "振替休日"), ("2026/7/20", "海の日"), ("2026/8/11", "山の日"),
    ("2026/9/21", "敬老の日"), ("2026/9/22", "国民の休日"), ("2026/9/23", "秋分の日"),
    ("2026/10/12", "スポーツの日"), ("2026/11/3", "文化の日"), ("2026/11/23", "勤労感謝の日"),
    ("2027/1/1", "元日"), ("2027/1/11", "成人の日"), ("2027/2/11", "建国記念の日"),
    ("2027/2/23", "天皇誕生日"), ("2027/3/22", "春分の日 振替"), ("2027/4/29", "昭和の日"),
    ("2027/5/3", "憲法記念日"), ("2027/5/4", "みどりの日"), ("2027/5/5", "こどもの日"),
    ("2027/7/19", "海の日"), ("2027/8/11", "山の日"), ("2027/9/20", "敬老の日"),
    ("2027/9/23", "秋分の日"), ("2027/10/11", "スポーツの日"), ("2027/11/3", "文化の日"),
    ("2027/11/23", "勤労感謝の日"),
]
for i, (d, name) in enumerate(holidays):
    y, m, dd = map(int, d.split("/"))
    c1 = ws.cell(row=5 + i, column=4, value=dt.date(y, m, dd))
    c1.number_format = "yyyy/m/d"
    c2 = ws.cell(row=5 + i, column=5, value=name)
    for c in (c1, c2):
        c.font = f_base
        c.border = b_all
        c.fill = fill_input
ws["D60"] = "※行 D5:D60 が判定対象"
ws["D60"].font = f_small

# 工種リスト
ws["H3"] = "工種リスト(編集可)"
ws["H3"].font = f_label
categories = ["仮設工事", "土工事・山留", "地業基礎工事", "躯体工事", "外装工事",
              "内装・設備工事", "外構工事", "検査・引渡", "マイルストーン", "その他"]
for i, cat in enumerate(categories):
    c = ws.cell(row=5 + i, column=8, value=cat)
    c.font = f_base
    c.border = b_all
    c.fill = fill_input

ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = 9
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws.print_area = "A1:J60"

wb.defined_names.add(DefinedName("工種リスト", attr_text="設定!$H$5:$H$14"))
wb.defined_names.add(DefinedName("祝日リスト", attr_text="設定!$D$5:$D$60"))

# ================================================================ 3. 工程マスタ
ws = wb.create_sheet("工程マスタ")
ws.sheet_properties.tabColor = "ED7D31"
widths = {"A": 5, "B": 14, "C": 30, "D": 16, "E": 11, "F": 11,
          "G": 6, "H": 8, "I": 5, "J": 24, "K": 10, "L": 10, "M": 10,
          "N": 6, "O": 6, "P": 6}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
for col in ("K", "L", "M"):
    ws.column_dimensions[col].hidden = True

ws["A1"] = "工程マスタ(作業の入力はこのシートだけ)"
ws["A1"].font = f_title
ws["A2"] = ("黄色列に入力 → 全体/月間/日割り/ネットワーク工程表へ自動反映。マイルストーンは開始日=終了日+MS列○。"
            "先行1〜3にはその作業より前に行う作業のNo(自分より小さいNoのみ)を入力。")
ws["A2"].font = f_small

headers = ["No", "工種", "作業名", "担当業者", "開始日", "終了日",
           "日数", "進捗率", "MS", "備考", "(開始)", "(終了)", "(実績境界)",
           "先行1", "先行2", "先行3"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = f_head
    c.fill = fill_head
    c.border = b_head
    c.alignment = al_c

# サンプルデータ (工期 2026/7/1〜2027/3/31、進捗は2026/7/20時点の例)
D = dt.date
tasks = [
    # (工種, 作業名, 業者, 開始, 終了, 進捗, MS, 備考)
    ("マイルストーン", "着工", "", D(2026, 7, 1), D(2026, 7, 1), None, "○", ""),
    ("マイルストーン", "地鎮祭・安全祈願祭", "", D(2026, 7, 4), D(2026, 7, 4), None, "○", ""),
    ("仮設工事", "仮囲い・ゲート設置", "○○仮設工業", D(2026, 7, 1), D(2026, 7, 10), 1.0, "", ""),
    ("仮設工事", "現場事務所・休憩所設置", "○○仮設工業", D(2026, 7, 6), D(2026, 7, 15), 1.0, "", ""),
    ("仮設工事", "仮設電気・給排水引込み", "△△電設", D(2026, 7, 13), D(2026, 7, 22), 0.6, "", ""),
    ("地業基礎工事", "杭工事", "□□基礎工業", D(2026, 7, 13), D(2026, 7, 31), 0.35, "", "既製杭"),
    ("土工事・山留", "山留工事", "▲▲土木", D(2026, 7, 27), D(2026, 8, 8), 0.0, "", ""),
    ("土工事・山留", "根切り・掘削", "▲▲土木", D(2026, 8, 3), D(2026, 8, 25), 0.0, "", ""),
    ("土工事・山留", "床付け・捨てコンクリート", "▲▲土木", D(2026, 8, 24), D(2026, 8, 29), 0.0, "", ""),
    ("地業基礎工事", "基礎配筋・型枠", "◇◇建設", D(2026, 8, 31), D(2026, 9, 12), 0.0, "", "配筋検査 9/9"),
    ("地業基礎工事", "基礎コンクリート打設", "◇◇建設", D(2026, 9, 14), D(2026, 9, 16), 0.0, "", ""),
    ("地業基礎工事", "養生・埋戻し", "▲▲土木", D(2026, 9, 17), D(2026, 9, 26), 0.0, "", ""),
    ("躯体工事", "鉄骨製作(工場)", "☆☆鉄工所", D(2026, 8, 3), D(2026, 9, 25), 0.0, "", "製品検査 9/18"),
    ("マイルストーン", "鉄骨建方開始", "", D(2026, 9, 28), D(2026, 9, 28), None, "○", ""),
    ("仮設工事", "外部足場組立", "○○仮設工業", D(2026, 9, 24), D(2026, 10, 6), 0.0, "", ""),
    ("躯体工事", "鉄骨建方", "☆☆鉄工所", D(2026, 9, 28), D(2026, 10, 23), 0.0, "", ""),
    ("マイルストーン", "上棟", "", D(2026, 10, 23), D(2026, 10, 23), None, "○", ""),
    ("躯体工事", "本締め・建入れ検査", "☆☆鉄工所", D(2026, 10, 26), D(2026, 11, 6), 0.0, "", ""),
    ("躯体工事", "デッキ・スタッド溶接", "☆☆鉄工所", D(2026, 10, 19), D(2026, 11, 13), 0.0, "", ""),
    ("躯体工事", "床コンクリート打設", "◇◇建設", D(2026, 11, 9), D(2026, 11, 27), 0.0, "", ""),
    ("外装工事", "屋根・屋上防水", "●●防水", D(2026, 11, 24), D(2026, 12, 18), 0.0, "", ""),
    ("外装工事", "外壁下地・胴縁", "■■外装", D(2026, 11, 16), D(2026, 12, 4), 0.0, "", ""),
    ("外装工事", "外壁パネル(ALC)取付", "■■外装", D(2026, 11, 30), D(2026, 12, 25), 0.0, "", ""),
    ("外装工事", "サッシ・外部建具取付", "▽▽サッシ", D(2026, 12, 14), D(2027, 1, 8), 0.0, "", ""),
    ("外装工事", "シール・外部防水", "●●防水", D(2027, 1, 5), D(2027, 1, 22), 0.0, "", ""),
    ("外装工事", "外壁塗装・仕上", "■■外装", D(2027, 1, 18), D(2027, 2, 5), 0.0, "", ""),
    ("マイルストーン", "外部足場解体", "", D(2027, 1, 25), D(2027, 1, 25), None, "○", ""),
    ("仮設工事", "外部足場解体", "○○仮設工業", D(2027, 1, 25), D(2027, 1, 30), 0.0, "", ""),
    ("内装・設備工事", "設備配管・配線(隠蔽部)", "△△電設・◎◎設備", D(2026, 11, 16), D(2027, 1, 16), 0.0, "", ""),
    ("内装・設備工事", "天井・壁LGS下地", "☆☆内装", D(2026, 12, 21), D(2027, 1, 29), 0.0, "", ""),
    ("内装・設備工事", "ボード貼り", "☆☆内装", D(2027, 1, 12), D(2027, 2, 12), 0.0, "", ""),
    ("マイルストーン", "受電", "", D(2027, 2, 2), D(2027, 2, 2), None, "○", ""),
    ("内装・設備工事", "塗装・クロス仕上", "☆☆内装", D(2027, 2, 8), D(2027, 3, 5), 0.0, "", ""),
    ("内装・設備工事", "床仕上げ", "☆☆内装", D(2027, 2, 22), D(2027, 3, 12), 0.0, "", ""),
    ("内装・設備工事", "照明・衛生器具取付", "△△電設・◎◎設備", D(2027, 2, 15), D(2027, 3, 10), 0.0, "", ""),
    ("内装・設備工事", "空調機器設置・試運転調整", "◎◎設備", D(2027, 2, 1), D(2027, 3, 16), 0.0, "", ""),
    ("外構工事", "外構土工・排水", "▲▲土木", D(2027, 2, 8), D(2027, 2, 26), 0.0, "", ""),
    ("外構工事", "舗装・植栽", "▲▲土木", D(2027, 3, 1), D(2027, 3, 19), 0.0, "", ""),
    ("検査・引渡", "社内検査・是正", "△△建設", D(2027, 3, 8), D(2027, 3, 17), 0.0, "", ""),
    ("マイルストーン", "消防・完了検査", "", D(2027, 3, 10), D(2027, 3, 10), None, "○", ""),
    ("検査・引渡", "美装クリーニング", "▼▼美装", D(2027, 3, 15), D(2027, 3, 24), 0.0, "", ""),
    ("マイルストーン", "施主検査", "", D(2027, 3, 18), D(2027, 3, 18), None, "○", ""),
    ("検査・引渡", "是正・手直し", "△△建設", D(2027, 3, 19), D(2027, 3, 29), 0.0, "", ""),
    ("仮設工事", "仮囲い・事務所撤去", "○○仮設工業", D(2027, 3, 22), D(2027, 3, 27), 0.0, "", ""),
    ("マイルストーン", "竣工・引渡し", "", D(2027, 3, 31), D(2027, 3, 31), None, "○", ""),
]


# ネットワーク工程表用の先行関係サンプル (作業No: [先行No])
PREDS = {2: [1], 3: [1], 4: [3], 5: [4], 6: [2, 3, 5], 7: [6], 8: [7], 9: [8],
         10: [9], 11: [10], 12: [11], 13: [2], 14: [12, 13], 15: [12],
         16: [14, 15], 17: [16], 18: [17], 19: [17], 20: [18, 19], 21: [20],
         22: [20], 23: [22], 24: [23], 25: [24], 26: [25], 27: [26], 28: [27],
         29: [20], 30: [21, 29], 31: [30], 32: [25], 33: [31], 34: [33],
         35: [31, 32], 36: [32], 37: [28], 38: [37], 39: [34, 35, 36],
         40: [39], 41: [39], 42: [40, 41], 43: [42], 44: [38, 43], 45: [43, 44]}
for r in range(MASTER_FIRST, MASTER_LAST + 1):
    i = r - MASTER_FIRST
    ws.cell(row=r, column=1, value=f'=IF($C{r}="","",ROW()-4)')
    if i < len(tasks):
        cat, name, sub, s, e, prog, ms, note = tasks[i]
        ws.cell(row=r, column=2, value=cat)
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=sub)
        ws.cell(row=r, column=5, value=s)
        ws.cell(row=r, column=6, value=e)
        if prog is not None:
            ws.cell(row=r, column=8, value=prog)
        if ms:
            ws.cell(row=r, column=9, value=ms)
        if note:
            ws.cell(row=r, column=10, value=note)
        for j, pv in enumerate(PREDS.get(i + 1, [])[:3]):
            ws.cell(row=r, column=14 + j, value=pv)
    ws.cell(row=r, column=7, value=f'=IF(OR($E{r}="",$F{r}=""),0,$F{r}-$E{r}+1)')
    ws.cell(row=r, column=11, value=f'=IF($E{r}="",0,$E{r})')
    ws.cell(row=r, column=12, value=f'=IF($F{r}="",0,$F{r})')
    ws.cell(row=r, column=13, value=f'=IF(OR($G{r}=0,$H{r}=""),$K{r},MIN($L{r}+1,$K{r}+$G{r}*$H{r}))')

style(ws, f"A{MASTER_FIRST}:P{MASTER_LAST}", font=f_base, border=b_all)
style(ws, f"N{MASTER_FIRST}:P{MASTER_LAST}", fill=fill_input, align=al_c)
style(ws, f"A{MASTER_FIRST}:A{MASTER_LAST}", fill=fill_calc, align=al_c)
style(ws, f"B{MASTER_FIRST}:F{MASTER_LAST}", fill=fill_input)
style(ws, f"E{MASTER_FIRST}:F{MASTER_LAST}", numfmt="yyyy/m/d", align=al_c)
style(ws, f"G{MASTER_FIRST}:G{MASTER_LAST}", fill=fill_calc, align=al_c, numfmt="0;-0;;@")
style(ws, f"H{MASTER_FIRST}:J{MASTER_LAST}", fill=fill_input)
style(ws, f"H{MASTER_FIRST}:H{MASTER_LAST}", numfmt="0%", align=al_c)
style(ws, f"I{MASTER_FIRST}:I{MASTER_LAST}", align=al_c)

dv_cat = DataValidation(type="list", formula1="=工種リスト", allow_blank=True)
dv_cat.add(f"B{MASTER_FIRST}:B{MASTER_LAST}")
ws.add_data_validation(dv_cat)
dv_ms = DataValidation(type="list", formula1='"○"', allow_blank=True)
dv_ms.add(f"I{MASTER_FIRST}:I{MASTER_LAST}")
ws.add_data_validation(dv_ms)
dv_prog = DataValidation(type="decimal", operator="between", formula1=0, formula2=1,
                         allow_blank=True, errorTitle="進捗率",
                         error="0〜1(0%〜100%)で入力してください")
dv_prog.add(f"H{MASTER_FIRST}:H{MASTER_LAST}")
ws.add_data_validation(dv_prog)
dv_pred = DataValidation(type="whole", operator="between", formula1=1, formula2=150,
                         allow_blank=True, errorTitle="先行作業",
                         error="先行作業のNoを入力してください(自分より小さいNoのみ)")
dv_pred.add(f"N{MASTER_FIRST}:P{MASTER_LAST}")
ws.add_data_validation(dv_pred)
# 不正な先行(自分以降のNo・空行参照)を赤で警告
ws.conditional_formatting.add(
    f"N{MASTER_FIRST}:P{MASTER_LAST}",
    FormulaRule(formula=[f'AND(N{MASTER_FIRST}<>"",OR(N{MASTER_FIRST}>=$A{MASTER_FIRST},'
                         f'INDEX($C$5:$C${MASTER_LAST},N{MASTER_FIRST})=""))'],
                stopIfTrue=True, fill=PatternFill("solid", bgColor="C00000"),
                font=Font(name=FONT, size=9, bold=True, color="FFFFFF")))

ws.freeze_panes = "A5"
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = 9
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws.print_area = f"A1:P{MASTER_LAST}"

# ================================================================ 4. 全体工程表
ws = wb.create_sheet("全体工程表")
ws.sheet_properties.tabColor = "1F4E79"
N_MONTH = 24
FIRST_COL = 8                     # H列
LAST_COL = FIRST_COL + N_MONTH - 1  # AE列 (31)
LAST_L = get_column_letter(LAST_COL)
G_FIRST = 6
G_LAST = G_FIRST + 99  # 105 (表示100行)
IDX_COL = 33   # AG: 抽出した工程マスタ行番号
BND_COL = 34   # AH: 実績境界
IDX_L, BND_L = "AG", "AH"

widths = {"A": 4, "B": 12, "C": 26, "D": 13, "E": 8, "F": 8, "G": 6}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
for c in range(FIRST_COL, LAST_COL + 1):
    ws.column_dimensions[get_column_letter(c)].width = 5.2
ws.column_dimensions[IDX_L].hidden = True
ws.column_dimensions[BND_L].hidden = True
ws.row_dimensions[3].hidden = True

ws["A1"] = '=設定!$B$3&"　全体工程表"'
ws["A1"].font = f_title
ws["A2"] = '="工期: "&TEXT(設定!$B$8,"yyyy/m/d")&" 〜 "&TEXT(設定!$B$9,"yyyy/m/d")&"　　発注者: "&設定!$B$4&"　　元請: "&設定!$B$5&"　　現場代理人: "&設定!$B$6'
ws["A2"].font = f_base
ws["L2"] = "現在の全体進捗率:"
ws["L2"].font = f_label
ws["N2"] = (f'=IF(SUMPRODUCT((工程マスタ!$K$5:$K${MASTER_LAST}>0)*(工程マスタ!$I$5:$I${MASTER_LAST}<>"○")*工程マスタ!$G$5:$G${MASTER_LAST})=0,0,'
            f'SUMPRODUCT((工程マスタ!$K$5:$K${MASTER_LAST}>0)*(工程マスタ!$I$5:$I${MASTER_LAST}<>"○")*工程マスタ!$G$5:$G${MASTER_LAST},工程マスタ!$H$5:$H${MASTER_LAST})'
            f'/SUMPRODUCT((工程マスタ!$K$5:$K${MASTER_LAST}>0)*(工程マスタ!$I$5:$I${MASTER_LAST}<>"○")*工程マスタ!$G$5:$G${MASTER_LAST}))')
ws["N2"].font = Font(name=FONT, size=11, bold=True, color=C_MS)
ws["N2"].number_format = "0.0%"
ws["S2"] = (f'=IF(SUMPRODUCT((工程マスタ!$K$5:$K${MASTER_LAST}>0)*1)>{G_LAST - G_FIRST + 1},'
            f'"※作業が{G_LAST - G_FIRST + 1}件を超えています。超過分は全体工程表に表示されません。","")')
ws["S2"].font = Font(name=FONT, size=9, bold=True, color=C_MS)

# 月ヘッダー (行3=月初シリアル(非表示), 行4=年, 行5=月)
for j in range(N_MONTH):
    col = FIRST_COL + j
    L = get_column_letter(col)
    prev = get_column_letter(col - 1)
    if j == 0:
        ws.cell(row=3, column=col, value="=設定!$B$12")
    else:
        ws.cell(row=3, column=col, value=f"=EDATE({prev}$3,1)")
    ws.cell(row=4, column=col,
            value=f'=IF(OR(COLUMN()={FIRST_COL},MONTH({L}$3)=1),YEAR({L}$3)&"年","")')
    ws.cell(row=5, column=col, value=f'=MONTH({L}$3)&"月"')

hdrs = ["No", "工種", "作業名", "担当業者", "開始", "終了", "進捗"]
for i, h in enumerate(hdrs, start=1):
    ws.cell(row=5, column=i, value=h)
style(ws, f"A4:{LAST_L}5", font=f_head, fill=fill_head, border=b_head, align=al_c)
style(ws, f"A4:G4", fill=fill_head)

# タスク行 (工程マスタから空行を詰めて自動抽出)
for r in range(G_FIRST, G_LAST + 1):
    k = r - G_FIRST + 1
    ws.cell(row=r, column=IDX_COL,
            value=ArrayFormula(f"{IDX_L}{r}",
                f"=IFERROR(SMALL(IF(工程マスタ!$K$5:$K${MASTER_LAST}>0,ROW(工程マスタ!$C$5:$C${MASTER_LAST})),{k}),0)"))
    ws.cell(row=r, column=BND_COL,
            value=f"=IF(${IDX_L}{r}=0,0,INDEX(工程マスタ!$M:$M,${IDX_L}{r}))")
    for ci, src in [(1, "$A:$A"), (2, "$B:$B"), (3, "$C:$C"), (4, "$D:$D"),
                    (5, "$E:$E"), (6, "$F:$F"), (7, "$H:$H")]:
        ws.cell(row=r, column=ci,
                value=(f'=IF(${IDX_L}{r}=0,"",IF(INDEX(工程マスタ!{src},${IDX_L}{r})="","",'
                       f'INDEX(工程マスタ!{src},${IDX_L}{r})))'))
    for c in range(FIRST_COL, LAST_COL + 1):
        L = get_column_letter(c)
        ws.cell(row=r, column=c,
                value=(f'=IF(OR(${IDX_L}{r}=0,{L}$3=""),"",'
                       f'IF(AND(INDEX(工程マスタ!$I:$I,${IDX_L}{r})="○",'
                       f'$E{r}>={L}$3,$E{r}<EDATE({L}$3,1)),"▼",""))'))

style(ws, f"A{G_FIRST}:G{G_LAST}", font=f_base, border=b_all)
style(ws, f"A{G_FIRST}:A{G_LAST}", align=al_c)
style(ws, f"E{G_FIRST}:F{G_LAST}", numfmt="m/d", align=al_c)
style(ws, f"G{G_FIRST}:G{G_LAST}", numfmt="0%", align=al_c)
style(ws, f"H{G_FIRST}:{LAST_L}{G_LAST}", font=f_ms, border=b_all, align=al_c)

# バー等の条件付き書式
ms_ex = f'INDEX(工程マスタ!$I:$I,${IDX_L}6)<>"○"'
ws.conditional_formatting.add(
    f"H{G_FIRST}:{LAST_L}{G_LAST}",
    FormulaRule(formula=[f'AND(${IDX_L}6>0,H$3<>"",$E6<EDATE(H$3,1),H$3<${BND_L}6,{ms_ex})'],
                stopIfTrue=True, fill=PatternFill("solid", bgColor=C_DONE)))
ws.conditional_formatting.add(
    f"H{G_FIRST}:{LAST_L}{G_LAST}",
    FormulaRule(formula=[f'AND(${IDX_L}6>0,H$3<>"",$E6<EDATE(H$3,1),$F6>=H$3,{ms_ex})'],
                stopIfTrue=True, fill=PatternFill("solid", bgColor=C_PLAN)))
ws.conditional_formatting.add(
    f"H{G_FIRST}:{LAST_L}{G_LAST}",
    FormulaRule(formula=['AND(H$3<>"",YEAR(H$3)=YEAR(TODAY()),MONTH(H$3)=MONTH(TODAY()))'],
                stopIfTrue=False, fill=PatternFill("solid", bgColor=C_TODAY)))

# 予定出来高曲線
CUM_ROW = G_LAST + 2
ws.cell(row=CUM_ROW, column=3, value="予定出来高(累計%)").font = f_label
base = (f'(工程マスタ!$K$5:$K${MASTER_LAST}>0)*(工程マスタ!$I$5:$I${MASTER_LAST}<>"○")')
den = f"SUMPRODUCT({base}*工程マスタ!$G$5:$G${MASTER_LAST})"
for c in range(FIRST_COL, LAST_COL + 1):
    L = get_column_letter(c)
    me = f"(EDATE({L}$3,1)-1)"
    num = (f"SUMPRODUCT({base}*(工程マスタ!$K$5:$K${MASTER_LAST}<={me})*"
           f"((工程マスタ!$L$5:$L${MASTER_LAST}<={me})*工程マスタ!$G$5:$G${MASTER_LAST}+"
           f"(工程マスタ!$L$5:$L${MASTER_LAST}>{me})*({me}-工程マスタ!$K$5:$K${MASTER_LAST}+1)))")
    ws.cell(row=CUM_ROW, column=c, value=f"=IF({den}=0,0,{num}/{den})")
style(ws, f"H{CUM_ROW}:{LAST_L}{CUM_ROW}", font=f_small, border=b_all,
      align=al_c, numfmt="0%")

# 実績出来高(黄色セルに毎月末の実績累計%を入力 → 青線でグラフに反映)
ACT_ROW = CUM_ROW + 1
ws.cell(row=ACT_ROW, column=3, value="実績出来高(累計%)※毎月末に入力").font = f_label
style(ws, f"H{ACT_ROW}:{LAST_L}{ACT_ROW}", font=f_small, fill=fill_input,
      border=b_all, align=al_c, numfmt="0%")
ws.cell(row=ACT_ROW, column=FIRST_COL, value=0.06)  # 入力例(2026年7月)

# 凡例
ws["U2"] = "凡例:"
ws["U2"].font = f_label
ws["V2"] = "予定"
ws["V2"].fill = PatternFill("solid", fgColor=C_PLAN)
ws["V2"].font = f_base
ws["V2"].alignment = al_c
ws["W2"] = "実績"
ws["W2"].fill = PatternFill("solid", fgColor=C_DONE)
ws["W2"].font = Font(name=FONT, size=9, color="FFFFFF")
ws["W2"].alignment = al_c
ws["X2"] = "▼ MS"
ws["X2"].font = f_ms

from openpyxl.drawing.line import LineProperties
from openpyxl.chart.series import SeriesLabel

chart = LineChart()
chart.title = "出来高曲線(赤=予定 / 青=実績)"
chart.style = 12
chart.height = 7
chart.width = 30
chart.y_axis.numFmt = "0%"
chart.y_axis.title = None
chart.x_axis.title = None
plan = Reference(ws, min_col=FIRST_COL, max_col=LAST_COL, min_row=CUM_ROW, max_row=CUM_ROW)
act = Reference(ws, min_col=FIRST_COL, max_col=LAST_COL, min_row=ACT_ROW, max_row=ACT_ROW)
cats = Reference(ws, min_col=FIRST_COL, max_col=LAST_COL, min_row=5, max_row=5)
chart.add_data(plan, from_rows=True, titles_from_data=False)
chart.add_data(act, from_rows=True, titles_from_data=False)
chart.set_categories(cats)
chart.series[0].tx = SeriesLabel(v="予定")
chart.series[0].smooth = True
chart.series[0].graphicalProperties.line = LineProperties(solidFill="C00000", w=22000)
chart.series[1].tx = SeriesLabel(v="実績")
chart.series[1].smooth = True
chart.series[1].graphicalProperties.line = LineProperties(solidFill="2F5597", w=22000)
chart.legend.position = "b"
ws.add_chart(chart, f"B{ACT_ROW + 2}")

ws.freeze_panes = "H6"
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = 8  # A3
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws.print_area = f"A1:{LAST_L}{ACT_ROW}"

# ================================================================ 5. 月間工程表
ws = wb.create_sheet("月間工程表")
ws.sheet_properties.tabColor = "2E7D32"
N_DAY = 31
FIRST_COL = 8                      # H
LAST_COL = FIRST_COL + N_DAY - 1   # AL (38)
LAST_L = get_column_letter(LAST_COL)
T_FIRST = 7
T_LAST = 56                        # 50行
IDX_COL, BND_COL = 40, 41          # AN, AO
IDX_L, BND_L = "AN", "AO"

widths = {"A": 4, "B": 12, "C": 26, "D": 13, "E": 8, "F": 8, "G": 6}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
for c in range(FIRST_COL, LAST_COL + 1):
    ws.column_dimensions[get_column_letter(c)].width = 3.4
ws.column_dimensions[IDX_L].hidden = True
ws.column_dimensions[BND_L].hidden = True
ws.row_dimensions[3].hidden = True

ws["A1"] = '=設定!$B$3&"　月間工程表（"&$B$2&"年"&$D$2&"月）"'
ws["A1"].font = f_title
ws["A2"] = "対象年"
ws["A2"].font = f_label
ws["B2"] = 2026
ws["B2"].fill = fill_input
ws["B2"].border = b_all
ws["B2"].font = f_base
ws["B2"].alignment = al_c
ws["C2"] = "対象月"
ws["C2"].font = f_label
ws["D2"] = 7
ws["D2"].fill = fill_input
ws["D2"].border = b_all
ws["D2"].font = f_base
ws["D2"].alignment = al_c
ws["E2"] = "← 黄色セルを変更すると自動で切替"
ws["E2"].font = f_small
ws[f"{IDX_L}1"] = "=DATE($B$2,$D$2,1)"
ws[f"{IDX_L}2"] = f"=EOMONTH(${IDX_L}$1,0)"

ws["N2"] = "凡例:"
ws["N2"].font = f_label
ws["O2"] = "予定"
ws["O2"].fill = PatternFill("solid", fgColor=C_PLAN)
ws["O2"].font = f_base
ws["O2"].alignment = al_c
ws["P2"] = "実績"
ws["P2"].fill = PatternFill("solid", fgColor=C_DONE)
ws["P2"].font = Font(name=FONT, size=9, color="FFFFFF")
ws["P2"].alignment = al_c
ws["Q2"] = "▼MS"
ws["Q2"].font = f_ms
ws["R2"] = "本日"
ws["R2"].fill = PatternFill("solid", fgColor="FFD966")
ws["R2"].font = f_base
ws["R2"].alignment = al_c

dv_y = DataValidation(type="whole", operator="between", formula1=2020, formula2=2045)
dv_y.add("B2")
ws.add_data_validation(dv_y)
dv_m = DataValidation(type="whole", operator="between", formula1=1, formula2=12)
dv_m.add("D2")
ws.add_data_validation(dv_m)

# 日ヘッダー
for j in range(N_DAY):
    col = FIRST_COL + j
    L = get_column_letter(col)
    ws.cell(row=3, column=col,
            value=f'=IF({j + 1}<=DAY(${IDX_L}$2),${IDX_L}$1+{j},"")')
    ws.cell(row=4, column=col, value=f'=IF({L}$3="","",DAY({L}$3))')
    ws.cell(row=5, column=col, value=f'=IF({L}$3="","",MID("日月火水木金土",WEEKDAY({L}$3),1))')

hdrs = ["No", "工種", "作業名", "担当業者", "開始", "終了", "進捗"]
for i, h in enumerate(hdrs, start=1):
    ws.cell(row=5, column=i, value=h)
style(ws, f"A4:{LAST_L}5", font=f_head, fill=fill_head, border=b_head, align=al_c)

# タスク行 (対象月に掛かる作業のみ自動抽出)
cond = (f"(工程マスタ!$K$5:$K${MASTER_LAST}>0)*(工程マスタ!$K$5:$K${MASTER_LAST}<=${IDX_L}$2)*(工程マスタ!$L$5:$L${MASTER_LAST}>=${IDX_L}$1)")
for r in range(T_FIRST, T_LAST + 1):
    k = r - T_FIRST + 1
    ws.cell(row=r, column=IDX_COL,
            value=ArrayFormula(f"{IDX_L}{r}",
                f"=IFERROR(SMALL(IF({cond},ROW(工程マスタ!$C$5:$C${MASTER_LAST})),{k}),0)"))
    ws.cell(row=r, column=BND_COL,
            value=f"=IF(${IDX_L}{r}=0,0,INDEX(工程マスタ!$M:$M,${IDX_L}{r}))")
    for ci, src in [(1, "$A:$A"), (2, "$B:$B"), (3, "$C:$C"), (4, "$D:$D"),
                    (5, "$E:$E"), (6, "$F:$F"), (7, "$H:$H")]:
        ws.cell(row=r, column=ci,
                value=(f'=IF(${IDX_L}{r}=0,"",IF(INDEX(工程マスタ!{src},${IDX_L}{r})="","",'
                       f'INDEX(工程マスタ!{src},${IDX_L}{r})))'))
    for c in range(FIRST_COL, LAST_COL + 1):
        L = get_column_letter(c)
        ws.cell(row=r, column=c,
                value=(f'=IF(OR(${IDX_L}{r}=0,{L}$3=""),"",'
                       f'IF(AND(INDEX(工程マスタ!$I:$I,${IDX_L}{r})="○",$E{r}={L}$3),"▼",""))'))

style(ws, f"A{T_FIRST}:G{T_LAST}", font=f_base, border=b_all)
style(ws, f"A{T_FIRST}:A{T_LAST}", align=al_c)
style(ws, f"E{T_FIRST}:F{T_LAST}", numfmt="m/d", align=al_c)
style(ws, f"G{T_FIRST}:G{T_LAST}", numfmt="0%", align=al_c)
style(ws, f"H{T_FIRST}:{LAST_L}{T_LAST}", font=f_ms, border=b_all, align=al_c)

# 稼働作業数
CNT_ROW = T_LAST + 1
ws.cell(row=CNT_ROW, column=3, value="稼働作業数").font = f_label
for c in range(FIRST_COL, LAST_COL + 1):
    L = get_column_letter(c)
    ws.cell(row=CNT_ROW, column=c,
            value=(f'=IF({L}$3="","",SUMPRODUCT((工程マスタ!$K$5:$K${MASTER_LAST}>0)*'
                   f'(工程マスタ!$I$5:$I${MASTER_LAST}<>"○")*(工程マスタ!$K$5:$K${MASTER_LAST}<={L}$3)*'
                   f'(工程マスタ!$L$5:$L${MASTER_LAST}>={L}$3)))'))
style(ws, f"C{CNT_ROW}:{LAST_L}{CNT_ROW}", font=f_small, border=b_all, align=al_c)

# 条件付き書式: 実績 → 予定 → 祝日/日曜/土曜
ms_ex = f'INDEX(工程マスタ!$I:$I,${IDX_L}7)<>"○"'
ws.conditional_formatting.add(
    f"H{T_FIRST}:{LAST_L}{T_LAST}",
    FormulaRule(formula=[f'AND(${IDX_L}7>0,H$3<>"",H$3>=$E7,H$3<${BND_L}7,{ms_ex})'],
                stopIfTrue=True, fill=PatternFill("solid", bgColor=C_DONE)))
ws.conditional_formatting.add(
    f"H{T_FIRST}:{LAST_L}{T_LAST}",
    FormulaRule(formula=[f'AND(${IDX_L}7>0,H$3<>"",H$3>=$E7,H$3<=$F7,{ms_ex})'],
                stopIfTrue=True, fill=PatternFill("solid", bgColor=C_PLAN)))
f_hd_black = Font(name=FONT, size=9, bold=True, color="000000")
for rng, hd in ((f"H4:{LAST_L}5", True), (f"H{T_FIRST}:{LAST_L}{CNT_ROW}", False)):
    fnt = f_hd_black if hd else None
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=['AND(H$3<>"",H$3=TODAY())'], stopIfTrue=True,
                         fill=PatternFill("solid", bgColor="FFD966"), font=fnt))
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=[f'AND(H$3<>"",COUNTIF(祝日リスト,H$3)>0)'],
                         stopIfTrue=True, fill=PatternFill("solid", bgColor=C_SUN), font=fnt))
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=['AND(H$3<>"",WEEKDAY(H$3)=1)'],
                         stopIfTrue=True, fill=PatternFill("solid", bgColor=C_SUN), font=fnt))
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=['AND(H$3<>"",WEEKDAY(H$3)=7)'],
                         stopIfTrue=True, fill=PatternFill("solid", bgColor=C_SAT), font=fnt))

ws.freeze_panes = "H7"
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = 8
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws.print_area = f"A1:{LAST_L}{CNT_ROW}"

# ================================================================ 6. 日割り工程表
ws = wb.create_sheet("日割り工程表")
ws.sheet_properties.tabColor = "00838F"
N_DAY = 14
FIRST_COL = 8                      # H
LAST_COL = FIRST_COL + N_DAY - 1   # U (21)
LAST_L = get_column_letter(LAST_COL)
NOTE_COL = 22                      # V: 備考
T_FIRST = 7
T_LAST = 56
IDX_COL, BND_COL = 24, 25          # X, Y
IDX_L, BND_L = "X", "Y"

widths = {"A": 4, "B": 12, "C": 26, "D": 13, "E": 8, "F": 8, "G": 6, "V": 20}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
for c in range(FIRST_COL, LAST_COL + 1):
    ws.column_dimensions[get_column_letter(c)].width = 6.5
ws.column_dimensions[IDX_L].hidden = True
ws.column_dimensions[BND_L].hidden = True
ws.row_dimensions[3].hidden = True

ws["A1"] = '=設定!$B$3&"　日割り工程表（"&TEXT($B$2,"yyyy/m/d")&" 〜 "&TEXT($B$2+13,"m/d")&"）"'
ws["A1"].font = f_title
ws["A2"] = "基準日"
ws["A2"].font = f_label
ws["B2"] = dt.date(2026, 7, 20)
ws["B2"].number_format = "yyyy/m/d"
ws["B2"].fill = fill_input
ws["B2"].border = b_all
ws["B2"].font = f_base
ws["B2"].alignment = al_c
ws["C2"] = "← 黄色セルに日付を入力すると14日間を表示"
ws["C2"].font = f_small

ws["N2"] = "凡例:"
ws["N2"].font = f_label
ws["O2"] = "予定"
ws["O2"].fill = PatternFill("solid", fgColor=C_PLAN)
ws["O2"].font = f_base
ws["O2"].alignment = al_c
ws["P2"] = "実績"
ws["P2"].fill = PatternFill("solid", fgColor=C_DONE)
ws["P2"].font = Font(name=FONT, size=9, color="FFFFFF")
ws["P2"].alignment = al_c
ws["Q2"] = "▼MS"
ws["Q2"].font = f_ms
ws["R2"] = "本日"
ws["R2"].fill = PatternFill("solid", fgColor="FFD966")
ws["R2"].font = f_base
ws["R2"].alignment = al_c

dv_d = DataValidation(type="date", operator="between",
                      formula1=dt.date(2020, 1, 1), formula2=dt.date(2045, 12, 31))
dv_d.add("B2")
ws.add_data_validation(dv_d)

for j in range(N_DAY):
    col = FIRST_COL + j
    L = get_column_letter(col)
    ws.cell(row=3, column=col, value=f"=$B$2+{j}")
    ws.cell(row=4, column=col, value=f'=TEXT({L}$3,"m/d")')
    ws.cell(row=5, column=col, value=f'=MID("日月火水木金土",WEEKDAY({L}$3),1)')

hdrs = ["No", "工種", "作業名", "担当業者", "開始", "終了", "進捗"]
for i, h in enumerate(hdrs, start=1):
    ws.cell(row=5, column=i, value=h)
ws.cell(row=5, column=NOTE_COL, value="備考")
style(ws, f"A4:V5", font=f_head, fill=fill_head, border=b_head, align=al_c)

cond = (f"(工程マスタ!$K$5:$K${MASTER_LAST}>0)*(工程マスタ!$K$5:$K${MASTER_LAST}<=$B$2+13)*(工程マスタ!$L$5:$L${MASTER_LAST}>=$B$2)")
for r in range(T_FIRST, T_LAST + 1):
    k = r - T_FIRST + 1
    ws.cell(row=r, column=IDX_COL,
            value=ArrayFormula(f"{IDX_L}{r}",
                f"=IFERROR(SMALL(IF({cond},ROW(工程マスタ!$C$5:$C${MASTER_LAST})),{k}),0)"))
    ws.cell(row=r, column=BND_COL,
            value=f"=IF(${IDX_L}{r}=0,0,INDEX(工程マスタ!$M:$M,${IDX_L}{r}))")
    for ci, src in [(1, "$A:$A"), (2, "$B:$B"), (3, "$C:$C"), (4, "$D:$D"),
                    (5, "$E:$E"), (6, "$F:$F"), (7, "$H:$H"), (NOTE_COL, "$J:$J")]:
        ws.cell(row=r, column=ci,
                value=(f'=IF(${IDX_L}{r}=0,"",IF(INDEX(工程マスタ!{src},${IDX_L}{r})="","",'
                       f'INDEX(工程マスタ!{src},${IDX_L}{r})))'))
    for c in range(FIRST_COL, LAST_COL + 1):
        L = get_column_letter(c)
        ws.cell(row=r, column=c,
                value=(f'=IF(${IDX_L}{r}=0,"",'
                       f'IF(AND(INDEX(工程マスタ!$I:$I,${IDX_L}{r})="○",$E{r}={L}$3),"▼",""))'))

style(ws, f"A{T_FIRST}:G{T_LAST}", font=f_base, border=b_all)
style(ws, f"V{T_FIRST}:V{T_LAST}", font=f_small, border=b_all)
style(ws, f"A{T_FIRST}:A{T_LAST}", align=al_c)
style(ws, f"E{T_FIRST}:F{T_LAST}", numfmt="m/d", align=al_c)
style(ws, f"G{T_FIRST}:G{T_LAST}", numfmt="0%", align=al_c)
style(ws, f"H{T_FIRST}:{LAST_L}{T_LAST}", font=f_ms, border=b_all, align=al_c)

CNT_ROW = T_LAST + 1
ws.cell(row=CNT_ROW, column=3, value="稼働作業数").font = f_label
for c in range(FIRST_COL, LAST_COL + 1):
    L = get_column_letter(c)
    ws.cell(row=CNT_ROW, column=c,
            value=(f'=SUMPRODUCT((工程マスタ!$K$5:$K${MASTER_LAST}>0)*(工程マスタ!$I$5:$I${MASTER_LAST}<>"○")*'
                   f'(工程マスタ!$K$5:$K${MASTER_LAST}<={L}$3)*(工程マスタ!$L$5:$L${MASTER_LAST}>={L}$3))'))
style(ws, f"C{CNT_ROW}:{LAST_L}{CNT_ROW}", font=f_small, border=b_all, align=al_c)

ms_ex = f'INDEX(工程マスタ!$I:$I,${IDX_L}7)<>"○"'
ws.conditional_formatting.add(
    f"H{T_FIRST}:{LAST_L}{T_LAST}",
    FormulaRule(formula=[f'AND(${IDX_L}7>0,H$3>=$E7,H$3<${BND_L}7,{ms_ex})'],
                stopIfTrue=True, fill=PatternFill("solid", bgColor=C_DONE)))
ws.conditional_formatting.add(
    f"H{T_FIRST}:{LAST_L}{T_LAST}",
    FormulaRule(formula=[f'AND(${IDX_L}7>0,H$3>=$E7,H$3<=$F7,{ms_ex})'],
                stopIfTrue=True, fill=PatternFill("solid", bgColor=C_PLAN)))
f_hd_black = Font(name=FONT, size=9, bold=True, color="000000")
for rng, hd in ((f"H4:{LAST_L}5", True), (f"H{T_FIRST}:{LAST_L}{CNT_ROW}", False)):
    fnt = f_hd_black if hd else None
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=['H$3=TODAY()'], stopIfTrue=True,
                         fill=PatternFill("solid", bgColor="FFD966"), font=fnt))
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=['COUNTIF(祝日リスト,H$3)>0'],
                         stopIfTrue=True, fill=PatternFill("solid", bgColor=C_SUN), font=fnt))
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=['WEEKDAY(H$3)=1'],
                         stopIfTrue=True, fill=PatternFill("solid", bgColor=C_SUN), font=fnt))
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=['WEEKDAY(H$3)=7'],
                         stopIfTrue=True, fill=PatternFill("solid", bgColor=C_SAT), font=fnt))

ws.freeze_panes = "H7"
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = 9  # A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws.print_area = f"A1:V{CNT_ROW}"


# ================================================================ 7. ネットワーク工程表
ws = wb.create_sheet("ネットワーク工程表")
ws.sheet_properties.tabColor = "7030A0"
N_WEEK = 52
FIRST_COL = 11                      # K
LAST_COL = FIRST_COL + N_WEEK - 1   # BJ (62)
LAST_L = get_column_letter(LAST_COL)
T_FIRST = 7
T_LAST = T_FIRST + (MASTER_LAST - MASTER_FIRST)  # 156
C_CP = "E06666"      # クリティカルバー(赤)
C_FLOAT = "FFE599"   # フロート(余裕幅)

widths = {"A": 4, "B": 24, "C": 5, "D": 8, "E": 7, "F": 7, "G": 7, "H": 7,
          "I": 5, "J": 4}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
for c in range(FIRST_COL, LAST_COL + 1):
    ws.column_dimensions[get_column_letter(c)].width = 2.8
ws.row_dimensions[3].hidden = True

ws["A1"] = '=設定!$B$3&"　ネットワーク工程表(CPM)"'
ws["A1"].font = f_title
ws["A2"] = "表示開始日"
ws["A2"].font = f_label
ws["B2"] = "=設定!$B$8"
ws["B2"].number_format = "yyyy/m/d"
ws["B2"].fill = fill_input
ws["B2"].border = b_all
ws["B2"].font = f_base
ws["B2"].alignment = al_c
ws["C2"] = "← 週単位で52週表示。先行のない作業は工期(自)から開始"
ws["C2"].font = f_small
ws["N2"] = "凡例:"
ws["N2"].font = f_label
ws["O2"] = "クリティカル"
ws.merge_cells("O2:R2")
ws["O2"].fill = PatternFill("solid", fgColor=C_CP)
ws["O2"].font = f_small
ws["O2"].alignment = al_c
ws["S2"] = "余裕あり"
ws.merge_cells("S2:V2")
ws["S2"].fill = PatternFill("solid", fgColor=C_PLAN)
ws["S2"].font = f_small
ws["S2"].alignment = al_c
ws["W2"] = "フロート"
ws.merge_cells("W2:Z2")
ws["W2"].fill = PatternFill("solid", fgColor=C_FLOAT)
ws["W2"].font = f_small
ws["W2"].alignment = al_c
ws["AA2"] = "▼MS"
ws.merge_cells("AA2:AC2")
ws["AA2"].font = f_ms
ws["AD2"] = "数字=作業No"
ws.merge_cells("AD2:AH2")
ws["AD2"].font = f_small

# 週ヘッダー (行3=週開始日シリアル(非表示), 行4=年月, 行5=日)
for j in range(N_WEEK):
    col = FIRST_COL + j
    L = get_column_letter(col)
    prev = get_column_letter(col - 1)
    if j == 0:
        ws.cell(row=3, column=col, value="=$B$2")
    else:
        ws.cell(row=3, column=col, value=f"={prev}$3+7")
    ws.cell(row=4, column=col,
            value=f'=IF(OR(COLUMN()={FIRST_COL},MONTH({L}$3)<>MONTH({L}$3-7)),TEXT({L}$3,"m月"),"")')
    ws.cell(row=5, column=col, value=f"=DAY({L}$3)")

hdrs = ["No", "作業名", "日数", "先行", "最早開始", "最早完了", "最遅開始", "最遅完了", "余裕", "CP"]
for i, h in enumerate(hdrs, start=1):
    ws.cell(row=6, column=i, value=h)
style(ws, f"A4:{LAST_L}6", font=f_head, fill=fill_head, border=b_head, align=al_c)
style(ws, f"A4:J5", fill=fill_head)

for r in range(T_FIRST, T_LAST + 1):
    m = r - 2          # 対応する工程マスタ行
    ws.cell(row=r, column=1, value=f'=IF(工程マスタ!$C{m}="","",工程マスタ!$A{m})')
    ws.cell(row=r, column=2, value=f'=IF($A{r}="","",工程マスタ!$C{m})')
    ws.cell(row=r, column=3, value=f'=IF($A{r}="","",MAX(1,工程マスタ!$G{m}))')
    ws.cell(row=r, column=4,
            value=f'=IF($A{r}="","",_xlfn.TEXTJOIN(",",TRUE,工程マスタ!$N{m}:$P{m}))')
    ws.cell(row=r, column=5,
            value=(f'=IF($A{r}="","",IF(COUNT(工程マスタ!$N{m}:$P{m})=0,設定!$B$8,'
                   f'MAX(IF(工程マスタ!$N{m}="",0,INDEX($F$7:$F${T_LAST},工程マスタ!$N{m})),'
                   f'IF(工程マスタ!$O{m}="",0,INDEX($F$7:$F${T_LAST},工程マスタ!$O{m})),'
                   f'IF(工程マスタ!$P{m}="",0,INDEX($F$7:$F${T_LAST},工程マスタ!$P{m})))+1))'))
    ws.cell(row=r, column=6, value=f'=IF($A{r}="","",$E{r}+$C{r}-1)')
    if m + 1 <= MASTER_LAST:
        cnt = (f'SUMPRODUCT((工程マスタ!$N${m + 1}:$N${MASTER_LAST}=$A{r})*1)+'
               f'SUMPRODUCT((工程マスタ!$O${m + 1}:$O${MASTER_LAST}=$A{r})*1)+'
               f'SUMPRODUCT((工程マスタ!$P${m + 1}:$P${MASTER_LAST}=$A{r})*1)')
        mn = (f'MIN(IF((工程マスタ!$N${m + 1}:$N${MASTER_LAST}=$A{r})+'
              f'(工程マスタ!$O${m + 1}:$O${MASTER_LAST}=$A{r})+'
              f'(工程マスタ!$P${m + 1}:$P${MASTER_LAST}=$A{r}),$G{r + 1}:$G${T_LAST}))-1')
        ws.cell(row=r, column=8, value=ArrayFormula(
            f"H{r}", f'=IF($A{r}="","",IF({cnt}=0,MAX($F$7:$F${T_LAST}),{mn}))'))
    else:
        ws.cell(row=r, column=8, value=f'=IF($A{r}="","",MAX($F$7:$F${T_LAST}))')
    ws.cell(row=r, column=7, value=f'=IF($A{r}="","",$H{r}-$C{r}+1)')
    ws.cell(row=r, column=9, value=f'=IF($A{r}="","",$G{r}-$E{r})')
    ws.cell(row=r, column=10, value=f'=IF($A{r}="","",IF($I{r}<=0,"●",""))')
    for c in range(FIRST_COL, LAST_COL + 1):
        L = get_column_letter(c)
        ws.cell(row=r, column=c,
                value=(f'=IF($A{r}="","",'
                       f'IF(AND(INDEX(工程マスタ!$I$5:$I${MASTER_LAST},$A{r})="○",'
                       f'$E{r}>={L}$3,$E{r}<={L}$3+6),"▼",'
                       f'IF(AND($E{r}>={L}$3,$E{r}<={L}$3+6),$A{r},"")))'))

style(ws, f"A{T_FIRST}:J{T_LAST}", font=f_base, border=b_all)
style(ws, f"A{T_FIRST}:A{T_LAST}", align=al_c)
style(ws, f"C{T_FIRST}:C{T_LAST}", align=al_c)
style(ws, f"D{T_FIRST}:D{T_LAST}", align=al_c)
style(ws, f"E{T_FIRST}:H{T_LAST}", numfmt="m/d", align=al_c)
style(ws, f"I{T_FIRST}:J{T_LAST}", align=al_c)
style(ws, f"K{T_FIRST}:{LAST_L}{T_LAST}",
      font=Font(name=FONT, size=7, color="44546A"), border=b_all, align=al_c)

# 条件付き書式
ws.conditional_formatting.add(
    f"A{T_FIRST}:J{T_LAST}",
    FormulaRule(formula=[f'AND($A{T_FIRST}<>"",$J{T_FIRST}="●")'],
                stopIfTrue=False, fill=PatternFill("solid", bgColor=C_SUN)))
ws.conditional_formatting.add(
    f"K{T_FIRST}:{LAST_L}{T_LAST}",
    FormulaRule(formula=[f'AND($A{T_FIRST}<>"",'
                         f'INDEX(工程マスタ!$I$5:$I${MASTER_LAST},$A{T_FIRST})="○",'
                         f'$E{T_FIRST}>=K$3,$E{T_FIRST}<=K$3+6)'],
                stopIfTrue=False,
                font=Font(name=FONT, size=7, bold=True, color=C_MS)))
ws.conditional_formatting.add(
    f"K{T_FIRST}:{LAST_L}{T_LAST}",
    FormulaRule(formula=[f'AND($A{T_FIRST}<>"",$J{T_FIRST}="●",'
                         f'$E{T_FIRST}<=K$3+6,$F{T_FIRST}>=K$3)'],
                stopIfTrue=False, fill=PatternFill("solid", bgColor=C_CP)))
ws.conditional_formatting.add(
    f"K{T_FIRST}:{LAST_L}{T_LAST}",
    FormulaRule(formula=[f'AND($A{T_FIRST}<>"",$J{T_FIRST}<>"●",'
                         f'$E{T_FIRST}<=K$3+6,$F{T_FIRST}>=K$3)'],
                stopIfTrue=True, fill=PatternFill("solid", bgColor=C_PLAN)))
ws.conditional_formatting.add(
    f"K{T_FIRST}:{LAST_L}{T_LAST}",
    FormulaRule(formula=[f'AND($A{T_FIRST}<>"",$H{T_FIRST}>$F{T_FIRST},'
                         f'$F{T_FIRST}+1<=K$3+6,$H{T_FIRST}>=K$3)'],
                stopIfTrue=True, fill=PatternFill("solid", bgColor=C_FLOAT)))
ws.conditional_formatting.add(
    f"K4:{LAST_L}{T_LAST}",
    FormulaRule(formula=['AND(K$3<=TODAY(),TODAY()<=K$3+6)'],
                stopIfTrue=False, fill=PatternFill("solid", bgColor=C_TODAY)))

ws.freeze_panes = "K7"
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = 8
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws.print_area = f"A1:{LAST_L}{T_FIRST + 99}"

# ================================================================ 保存
out = "工事工程表アプリ.xlsx"
wb.save(out)
print(f"saved: {out}")
