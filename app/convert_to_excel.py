"""
CSVファイルをPower Apps用Excelワークブック（.xlsx）に変換するスクリプト
--------------------------------------------------
使い方:
    pip install openpyxl
    python app/convert_to_excel.py

生成されるファイル:
    data/NanbuIchibaDelivery.xlsx
        - Categories シート
        - Stores シート
        - MenuItems シート
        - Orders シート
        - OrderItems シート

Power Apps での使い方:
    1. 生成した .xlsx を OneDrive / SharePoint にアップロード
    2. Power Apps Studio でデータソース追加 → Excel Online (Business)
    3. 各シートのテーブル名を選択して接続
"""

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent / "data"
OUTPUT_FILE = DATA_DIR / "NanbuIchibaDelivery.xlsx"

HEADER_FILL = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Meiryo UI", size=11)
CELL_FONT   = Font(name="Meiryo UI", size=10)
BORDER_SIDE = Side(style="thin", color="CCCCCC")
CELL_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)

CSV_FILES = [
    ("Categories", "categories_table"),
    ("Stores",     "stores_table"),
    ("MenuItems",  "menu_items_table"),
    ("Orders",     "orders_table"),
    ("OrderItems", "order_items_table"),
]


def load_csv(name: str) -> list[dict]:
    path = DATA_DIR / f"{name}.csv"
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_sheet(wb: openpyxl.Workbook, sheet_name: str, table_name: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(title=sheet_name)
    if not rows:
        return

    headers = list(rows[0].keys())
    # ヘッダー行を書き込む
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER

    # データ行を書き込む
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(headers, start=1):
            value = row.get(key, "")
            # 型変換
            if value.upper() in ("TRUE", "FALSE"):
                value = value.upper() == "TRUE"
            else:
                try:
                    value = int(value)
                except ValueError:
                    try:
                        value = float(value)
                    except ValueError:
                        pass

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = CELL_BORDER

    # 列幅を自動調整
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(header)),
            *[len(str(r.get(header, ""))) for r in rows]
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 50)

    ws.row_dimensions[1].height = 24

    # Excel テーブルとして登録（Power Apps で認識される）
    last_col = get_column_letter(len(headers))
    last_row = len(rows) + 1
    table_ref = f"A1:{last_col}{last_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TABLE_STYLE
    ws.add_table(table)

    ws.freeze_panes = "A2"


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    for sheet_name, table_name in CSV_FILES:
        print(f"  シート作成: {sheet_name} ...")
        rows = load_csv(sheet_name)
        write_sheet(wb, sheet_name, table_name, rows)
        print(f"    → {len(rows)} 行")

    wb.save(OUTPUT_FILE)
    print(f"\n✅ 完成: {OUTPUT_FILE}")
    print("\n--- Power Apps での接続手順 ---")
    print("1. NanbuIchibaDelivery.xlsx を OneDrive または SharePoint にアップロード")
    print("2. Power Apps Studio → データ → データの追加")
    print("3. 'Excel Online (Business)' または 'OneDrive for Business' を選択")
    print("4. ファイルを選択 → 以下のテーブルをすべてチェック:")
    for _, table_name in CSV_FILES:
        print(f"     ✔ {table_name}")
    print("5. 接続 → powerapps/Src/ の YAML を参考にアプリを構築")


if __name__ == "__main__":
    print("🍱 南部市場 出前アプリ - Excel 変換スクリプト\n")
    main()
